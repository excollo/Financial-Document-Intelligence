"""
Excel Builder — generates a multi-sheet spreadsheet report 
from the extracted section results.
"""
import io
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from typing import List, Dict, Any
from app.core.logging import get_logger

logger = get_logger(__name__)


class ExcelBuilder:
    """
    Constructs a professional DRHP Analysis report in Excel format.
    One sheet per section + a Summary sheet.
    """

    def build(self, job_details: Dict[str, Any], section_results: List[Dict[str, Any]]) -> bytes:
        """
        Create the Excel workbook in memory.
        
        Args:
            job_details: Metadata about the job (company name, date, etc)
            section_results: List of result documents from MongoDB
            
        Returns:
            Bytes of the generated .xlsx file
        """
        wb = openpyxl.Workbook()
        
        # 1. Summary Sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Header Styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        summary_sheet["A1"] = "DRHP / RHP Intelligence Report - Summary"
        summary_sheet["A1"].font = header_font
        summary_sheet.merge_cells("A1:C1")
        summary_sheet["A1"].fill = header_fill
        
        summary_sheet["A3"] = "Company Name"
        summary_sheet["B3"] = job_details.get("document_name", "N/A")
        
        summary_sheet["A4"] = "Job ID"
        summary_sheet["B4"] = job_details.get("job_id", "N/A")
        
        summary_sheet["A5"] = "Generated At"
        summary_sheet["B5"] = job_details.get("created_at", "N/A")
        
        # 2. Detailed Sheets (One per Section)
        for res in section_results:
            section_id = res.get("section_id", "Unknown")
            # Excel sheet names limited to 31 chars
            title = section_id[:31]
            try:
                sheet = wb.create_sheet(title=title)
            except Exception:
                # Handle duplicate titles if necessary
                sheet = wb.create_sheet(title=f"{title[:28]}_{res.get('_id')[-2:]}")
                
            raw_json = res.get("raw_json", {})
            
            # Simple Column 1: Field, Column 2: Value
            sheet["A1"] = f"Extraction Results: {section_id.replace('_', ' ').capitalize()}"
            sheet["A1"].font = Font(bold=True, size=12)
            
            row = 3
            for key, val in raw_json.items():
                if key == "_markdown": continue
                
                sheet.cell(row=row, column=1, value=key.replace("_", " ").title())
                sheet.cell(row=row, column=1).font = Font(bold=True)
                
                # Check if value is a list (table data)
                if isinstance(val, list):
                    sheet.cell(row=row, column=2, value="[List Data / Table]")
                    # For a real implementation, we could expand the list into a mini-table here.
                else:
                    sheet.cell(row=row, column=2, value=str(val) if val is not None else "N/A")
                row += 1
                
            # Auto-size columns for better readability
            for col in ['A', 'B']:
                sheet.column_dimensions[col].width = 30

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
